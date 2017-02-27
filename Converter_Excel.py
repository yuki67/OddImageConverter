import openpyxl
import Converter


def convert(array, size, filename):
    ConverterExcel().convert(array, size, filename)


class ConverterExcel(Converter.Converter):
    """ エクセルでお絵かきするためのクラス """

    def __init__(self, cell_size=10):
        super().__init__()
        self.cell_size = cell_size
        self.x = self.y = 0
        self.filename = None
        self.file = None

    def open(self, color_array, size, filename):
        """ 初期設定を行う """
        self.filename = filename + ".xlsx"
        self.file = openpyxl.Workbook()
        sheet = self.file.active
        d = max(1, max(len(color_array), len(color_array[0])) / size)
        for y in range(int(len(color_array) / d) + 1):
            # pylint: disable=maybe-no-member
            sheet.row_dimensions[y + 1].height = self.cell_size / 2
            for x in range(int(len(color_array[0]) / d)):
                letter = openpyxl.utils.get_column_letter(x + 1)
                # 11.07は実測値
                sheet.column_dimensions[letter].width = self.cell_size / 11.07
            # pylint: enable=maybe-no-member

    def put_pixel(self, color):
        """ (x, y)にcolorを描画する """
        key = openpyxl.utils.get_column_letter(self.y + 1) + str(self.x + 1)
        self.file.active[key].fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=self.x11_from_rgb(color))
        self.x += 1

    def new_line(self):
        """ 行を改める """
        self.x = 0
        self.y += 1

    def close(self):
        """ ファイルを閉じる """
        self.file.save(filename=self.filename)

    @staticmethod
    def x11_from_rgb(rgb):
        """
        色のRGB表記(R, G, B)をX11表記(#ARGB)に直す
        aはffで固定
        """
        return "ff" + hex(rgb[0])[2:].rjust(2, '0') + hex(rgb[1])[2:].rjust(2, '0') + hex(rgb[2])[2:].rjust(2, '0')
